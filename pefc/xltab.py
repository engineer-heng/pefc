# Python Engineering Foundation Class Library (pefc)
# xltab Library
# Copyright(C) 2017-2025 Heng Swee Ngee
#
# Released under the MIT License - https://opensource.org/licenses/MIT
#

"""
xltab Library - Tabulate data from MS Excel forms. MS Excel workbooks
typically contains data arranged in form layout used for data entry.
Such forms are useful for people but unsuitable for data analysis.
Use the classes here to tabulate data from multiple forms in MS Excel
format by using the class method tab_form.
Class methods here return data in either a single value, a list or
dict that can be easily converted into a database table or
pandas DataFrame to be processed by Apps for data mining and
data analysis.
"""

from abc import ABC, abstractmethod
import copy
import datetime
import math
import openpyxl
import openpyxl.utils.cell
import xlrd
import xlwings


# LIBRARY DESIGN DECISIONS
# D02 Added abstract class PyXLDataLink - Heng Swee Ngee 2017-08-27
# PyXLDataLink is an abstract base class for getting data from MS Excel.
# OXLDataLink uses openpyxl while XLDataLink uses xlrd.
# PEfc's XLDataLink is not the same as JEfc's XLDataLink because python
# libraries openpyxl, xlwings and xlrd provides better functionality
# with little coding.
# OXLDataLink, XLDataLink and XLWDataLink are more flexible and easier
# to use than the JEfc equivalent classes
# Did not check the correctness of cell reference or cell address but
# leave it to the underlying library. (DRY)
# Decide to follow pandas docstring guidelines


class PyXLDatalink(ABC):
    """
    Abstract base class for OXLDataLink, XLDataLink and
    XLWDatalink sub-classes.
    """

    # Cell reference indicators
    _sheet_ind = '!'
    _fixed_ind = '$'

    def __init__(self, xlfname, xlsheet=0):  # default first sheet
        self._file_name = xlfname
        self._sheet_name = xlsheet
        self._wbook = None
        self._wsheet = None

    @property
    @abstractmethod
    def sheet(self):
        pass

    @sheet.setter
    @abstractmethod
    def sheet(self, sheet_spec):
        pass

    @abstractmethod
    def __str__(self):
        pass

    @abstractmethod
    def __repr__(self):
        pass

    @property
    def workbook(self):
        """ Returns the workbook object
        """
        return self._wbook

    @property
    def worksheet(self):
        """ Returns the worksheet object
        """
        return self._wsheet

    @property
    def file_name(self):
        """ Returns the MS Excel file name
        """
        return self._file_name

    @property
    def sheet_name(self):
        """ Returns the worksheet name
        """
        return self._sheet_name

    @abstractmethod
    def range_value(self, cell_ref, dates=datetime.date,
                    ndim=1, empty=math.nan) -> list:
        pass

    @abstractmethod
    def raw_range_value(self, cell_ref):
        pass

    @abstractmethod
    def row_values(self, xl_row_no: int, opt_date=datetime.date,
                   opt_empty=math.nan):
        pass

    @abstractmethod
    def col_values(self, xl_col_str: str, opt_date=datetime.date,
                   opt_empty=math.nan):
        pass

    def tab_form(self, cell_refs, opt_date=datetime.date, opt_empty=math.nan):
        """
        This method extracts data from an MS Excel workbook using
        the cell references. It process data type automatically.
        If cell errors are need to be analyzed then use
        raw_range_value method.

        Parameters
        ---------
        cell_refs: str
            * A cell reference e.g 'A7', 'A7:G10';
            * A list of cell references e.g. ['D23', None, 'E23'],
            * A dict of cell references e.g.
            {'Name': 'B12',
             'Age':'B15',
             'Height': 'C20'}
            * If a None is placed in a cell reference list or dict it is
            treated as a placeholder.
        opt_date : datetime.datetime or default datetime.date
            e.g. datetime.date returns dates in YYYY-MM-DD
        opt_empty : str or default math.nan
            Return this specified value if cell is blank, empty or error.
            e.g 'NA', '' or nan

        Returns
        -------
        A single value, list of values or dict of values depending
        on cell_refs parameter

        Examples
        --------
        >>> xldl = XLDataLink(r'./samples/xltest.xlsx')
        >>> # Or xldl = OXLDataLink(r'./samples/xltest.xlsx')
        >>> # XLWDataLink need xlwings.App() and 'with' syntax e.g.
        >>> # with XLWDataLink(xlapp, r'./samples/xltest.xlsx') as xldl:
        >>> cell_ls = ['A13', 'B13', 'C13', 'D13', None, 'E13',
        >>> 'F13', 'G13', 'H13', 'I13', 'J13', 'A34:I34']
        >>> data_ls = xldl.tab_form(cell_ls, opt_empty='NA')
        >>> # return value is
        >>> ['NA', 32.7888, 63.0, datetime.date(2014, 9, 24), nan,
        >>> 'a string', True, nan, 2065.6944, datetime.date(2017, 9, 1),
        >>> datetime.date(1902, 12, 7), [1.0, 32.7888, 20.0, 'C',
        >>> 'String test C', False, 1.0, 655.7760000000001, 1.0]]
        >>> xldl.sheet = 'Sheet2'
        >>> cref_dc = {'Name': 'B12', 'Age': 'B15', 'Height': 'C20'}
        >>> values_dc = xldl.tab_form(cref_dc)
        >>> # return value is
        >>> {'Name': 'John Doe', 'Age': 33.0, 'Height': 168.0}
        """

        # check if xw_sheet is valid
        if self._wsheet is None:
            print('Sheet does not exist or sheet specified is invalid')
            return None
        # Data Extraction
        if isinstance(cell_refs, list):
            value_ls = []
            # gets data from specified sheet
            for cell_ref in cell_refs:
                # put None in reference to allocate a blank cell to
                # fill with values later
                if cell_ref is None:
                    value_ls.append(math.nan)
                else:
                    value = self.range_value(
                        cell_ref, dates=opt_date, empty=opt_empty)
                    value_ls.append(value)
            # returns a list so that the user can modify it
            return value_ls
        elif isinstance(cell_refs, str):
            if ':' in cell_refs:
                value_ls = self.range_value(
                    cell_refs, dates=opt_date, ndim=2, empty=opt_empty)
                if len(value_ls) == 1:  # 'A68:H68' data row returns a list
                    # change to list instead of a 2-dim list
                    value_ls = value_ls[0]
                return value_ls  # return 2-dim list or a list
            elif cell_refs is None:  # unlikely have None in a single cell ref
                return math.nan
            else:  # single cell reference e.g. 'A8' return a single value
                return self.range_value(cell_refs, dates=opt_date,
                                        empty=opt_empty)
        elif isinstance(cell_refs, dict):
            value_dc = copy.deepcopy(cell_refs)
            # replace cell_ref in dict key values with the values obtained
            # from the cell e.g.
            # dict of {'qty': 'A8', 'total': 'C7'} to
            # {'qty': 80, 'total': 1000}
            for k, val in value_dc.items():
                # put None in reference to allocate a blank cell
                # to fill with values later
                if val is None:
                    value_dc[k] = math.nan
                else:
                    value_dc[k] = self.range_value(
                        val, dates=opt_date, empty=opt_empty)
            return value_dc  # returns dict
        else:
            return None

    @abstractmethod
    def set_cell_value(self, cell_ref, values):
        pass

    @abstractmethod
    def save(self, fname=None):
        pass

    def fill_form(self, cell_refs_val):
        """
        Fill the form in the worksheet with data.
        cell_refs_val: tuple list or dict.
        e.g. (('A7', 68035.75), ('C7', 'Red'), ('D7', True})
        or  {'A7': 68035.75,
             'C7': 'Red',
             'D7': True}
        """
        if isinstance(cell_refs_val, list):
            # tuple list. (('A7', 68035.75), ('C7', 'Red'), ('D7', True})
            for tup in cell_refs_val:
                self.set_cell_value(*tup)
        elif isinstance(cell_refs_val, tuple):  # e.g. ('A8', 811)
            self.set_cell_value(*cell_refs_val)
        # e.g. dict of {'A7': 68035.75, 'C7': 'Red', 'D7': True}
        elif isinstance(cell_refs_val, dict):
            for k, val in cell_refs_val.items():
                if k is None:
                    return None
                else:
                    if val is None:
                        # Will blank the cell in set_cell_value
                        self.set_cell_value(k, math.nan)
                    else:
                        self.set_cell_value(k, val)

    # Port Andrew Khan's java functions from
    # JExcelApi CellReferenceHelper class
    @classmethod
    def cref_num_index(cls, crs):
        """
        Find the position of the first number in a cell reference
        e.g. Sheet1!A10 returns 8

        crs: cell reference str
        return: position of first number
        """
        pos = crs.rfind(cls._sheet_ind) + 1
        for char in crs[pos:]:
            if 48 <= ord(char) <= 57:  # '0' = 48 and '9' = 57
                break
            else:
                pos += 1
        return pos

    @classmethod
    def is_cref_col_abs(cls, crs):
        """
        Checks if column is an absolute column address
        e.g. $F35 will return True

        crs: cell reference str
        return: bool
        """
        pos = crs.rfind(cls._sheet_ind) + 1  # in case of 'Sheet1!$A10'
        return crs[pos] == cls._fixed_ind

    @classmethod
    def is_cref_row_abs(cls, crs):
        """
        Checks if row is an absolute row address e.g. F$35 will return True

        crs: cell reference str
        return: bool
        """
        return crs[cls.cref_num_index(crs) - 1] == cls._fixed_ind

    @classmethod
    def cref_coord(cls, crs):
        pass


class OXLDataLink(PyXLDatalink):
    """
    OXLDataLink uses openpyxl therefore row and column numbers
    are 1-based. Can only be used for MS Excel 2007 and above
    file formats. Primary be used for data extraction however
    fill_form and save method is supported.
    """

    def __init__(self, xlfname, xlsheet=0):  # default first sheet
        super().__init__(xlfname, xlsheet)
        self._wbook = openpyxl.load_workbook(xlfname, data_only=True)
        self._file_name = xlfname
        self.sheet = xlsheet

    @property
    def sheet(self):
        return self._sheet_name

    @sheet.setter
    def sheet(self, sheet_spec):
        if isinstance(sheet_spec, int):
            self._sheet_name = self._wbook.sheetnames[sheet_spec]
        elif isinstance(sheet_spec, str):
            self._sheet_name = sheet_spec
        self._wsheet = self._wbook[self._sheet_name]

    def __str__(self):
        return f'{self.__class__.__name__}: Book={self._file_name},\
 Sheet={self._sheet_name}'

    def __repr__(self):
        return f'{self.__class__.__name__}\
({self._file_name!r}, {self._sheet_name!r})'

    @staticmethod
    def cell_value(cell, dates=datetime.date, empty=math.nan):
        value = cell.value
        if value is None:
            return empty
        if cell.is_date:
            # process date options
            # Assumes that openpyxl handles the excel base date
            # differences with Mac
            # cell.base_date = 2415018.5 is offset for python datetime
            if dates is datetime.datetime:
                return value
            elif dates is datetime.date:
                # value can be a datetime.datetime or datetime.date
                if isinstance(value, datetime.datetime):
                    return value.date()
                else:
                    return value
        elif cell.data_type == cell.TYPE_ERROR:
            return math.nan
        else:
            return value

    def range_value(self, cell_ref, dates=datetime.date, ndim=1,
                    empty=math.nan):
        if cell_ref is None:
            return math.nan  # error return Nan for robustness
        elif isinstance(cell_ref, list):
            return math.nan  # error return Nan for robustness
        elif ':' in cell_ref:  # range type of cell_ref
            cells = self._wsheet[cell_ref]
            if ndim == 1:
                return [self.cell_value(cell, dates, empty)
                        for row in cells for cell in row]
            else:
                # return 2-dim list or 1-dim list
                value_ls = []
                for row in cells:
                    row_ls = []
                    for cell in row:
                        row_ls.append(self.cell_value(cell, dates, empty))
                    value_ls.append(row_ls)
        else:  # single cell reference
            value_ls = self.cell_value(self._wsheet[cell_ref], dates, empty)
        return value_ls

    def raw_range_value(self, cell_ref):
        if cell_ref is None:
            return None  # return None to indicate error unlike range_value
        elif isinstance(cell_ref, list):
            return None  # return None to indicate error unlike range_value
        elif ':' in cell_ref:  # range type of cell_ref
            cells = self._wsheet[cell_ref]
            value_ls = []
            for row in cells:
                row_ls = []
                for cell in row:
                    row_ls.append(cell.value)
                value_ls.append(row_ls)
            if len(value_ls) == 1:
                value_ls = value_ls[0]
        else:  # single cell ref
            cell = self._wsheet[cell_ref]
            value_ls = cell.value
        return value_ls

    def row_values(self, xl_row_no: int, opt_date=datetime.date,
                   opt_empty=math.nan):
        """
        Returns the cell values in the specified row number.

        Parameters
        ----------
        xl_row_no : int
            MS Excel row number. 1-based index
        opt_date : datetime.datetime or default datetime.date
        opt_empty : str or default math.nan
            Sets empty or blank cell return value e.g 'NA', '' or nan

        Returns
        -------
            list of cell values
        """
        rowv = self._wsheet[xl_row_no]
        return [self.cell_value(cell, opt_date, opt_empty) for cell in rowv]

    def col_values(self, xl_col_str: str, opt_date=datetime.date,
                   opt_empty=math.nan):
        """
        Returns the cell values in the specified column address.

        Parameters
        ----------
        xl_col_str : str
            MS Excel column address e.g. 'A'
        opt_date : datetime.datetime or default datetime.date
        opt_empty : str or default math.nan
            Sets empty or blank cell return value e.g 'NA', '' or nan

        Returns
        -------
            list of cell values
        """
        colv = self._wsheet[xl_col_str]
        return [self.cell_value(cell, opt_date, opt_empty) for cell in colv]

    def set_cell_value(self, cell_ref, values):
        if cell_ref is None:
            return None  # since no valid cell address do nothing
        elif isinstance(cell_ref, list):
            return None   # cell_ref should not be a list here
        elif ':' in cell_ref:  # range type of cell_ref
            cells = self._wsheet[cell_ref]
            # convert to 1-dim tuples
            cells_tu = (cell for row in cells for cell in row)
            values_tu = (value for row in values for value in row)
            tu_ls = zip(cells_tu, values_tu)
            for cell, value in tu_ls:
                cell.value = value
        else:   # single cell address
            cell = self._wsheet[cell_ref]
            cell.value = values

    def save(self, fname=None):
        if fname is None:
            self._wbook.save(self._file_name)
        else:
            self._wbook.save(fname)

    @classmethod
    def cref_coord(cls, crs):
        """
        Returns the row and column number of a cell reference
        e.g. 'AZ5' returns (5, 52)
        openpyxl coordinates is 1-based index unlike xlrd which
        is 0-based. Returned row and column numbers are 1-based.

        crs: cell reference
        return: returns (row_no, col_no) tuple
        """
        return openpyxl.utils.cell.coordinate_to_tuple(crs)


class XLDataLink(PyXLDatalink):
    """
    XLDataLink uses xlrd therefore row and column numbers are
    0-based unlike OXLDataLink and XLWDataLink to maintain
    compatibility with xlrd library.

    XLDatalink are used only to extract data. fill_form
    and save method not supported. Both xls and xlsx format
    are supported.

    XLDataLink is a good extension of the xlrd library since
    it makes use of cell references instead of the row and
    column numbers. XLDataLink provides conversion functions
    for cell references to row and column number format.
    """

    def __init__(self, xlfname, xlsheet=0):  # default first sheet
        super().__init__(xlfname, xlsheet)
        # formatting_info=True not implemented for xlsx
        self._wbook = xlrd.open_workbook(xlfname)
        self.sheet = xlsheet
        # 0: 1900-based, 1: 1904-based.
        self._dmode = self._wsheet.book.datemode

    @property
    def sheet(self):
        return self._sheet_name

    @sheet.setter
    def sheet(self, sheet_spec):
        if isinstance(sheet_spec, int):
            self._wsheet = self._wbook.sheet_by_index(sheet_spec)
            self._sheet_name = self._wsheet.name
        elif isinstance(sheet_spec, str):
            self._wsheet = self._wbook.sheet_by_name(sheet_spec)
            self._sheet_name = self._wsheet.name

    def __str__(self):
        return f'{self.__class__.__name__}: Book={self._file_name},\
 Sheet={self._sheet_name}'

    def __repr__(self):
        return f'{self.__class__.__name__}\
({self._file_name!r}, {self._sheet_name!r})'

    def range_value(self, cell_ref, dates=datetime.date, ndim=1,
                    empty=math.nan):
        """
        Returns cell value from a single cell reference or cell values
        from a cell range reference.
        If cell error returns math.nan to indicate a cell error.
        If cell_ref is None or a list it returns math.nan as placeholder

        Parameters
        ----------
        cell_ref : str
            Single cell reference e.g 'H82' or cell range e.g. 'B5:E36'
        dates : datetime.datetime or default datetime.date
        ndim: int
            default 1 for 1D list or 2 for 2D list
        date : datetime.datetime or default datetime.date
        empty : str or default math.nan
            Sets empty or blank cell return value e.g 'NA', '' or nan

        Returns
        -------
        Single cell value or list of cell values:
        """
        # handles a single cell ref or a cell range only
        if cell_ref is None:
            return math.nan  # error return Nan for robustness
        elif isinstance(cell_ref, list):
            return math.nan  # error return Nan for robustness
        elif ':' in cell_ref:  # range type of cell_ref
            crg_tu = cell_ref.partition(':')
            rcs = self.cref_coord(crg_tu[0])
            rce = self.cref_coord(crg_tu[2])
            return self.cell_range_value(*rcs, *rce, dates, ndim, empty)
        # normal cell_ref - no check on validity of cell_ref. Leave it to xlrd
        rc_ = self.cref_coord(cell_ref)
        return self.cell_value(*rc_, dates, empty)

    def raw_range_value(self, cell_ref):
        """
        Get the raw cell value or range of raw cell values as is without
        cell type modifications.
        Useful for checking cell errors or for debugging errors in a sheet.

        cell_ref: str
            Single cell reference e.g 'H82' or cell range e.g. 'B5:E36'

        Returns
        -------
        Single raw cell value or list of raw cell values
        """
        # handles a single cell ref or a cell range only
        if cell_ref is None:
            return None  # return None to indicate error unlike range_value
        elif isinstance(cell_ref, list):
            return None  # return None to indicate error unlike range_value
        elif ':' in cell_ref:  # range type of cell_ref
            crg_tu = cell_ref.partition(':')
            rcs = self.cref_coord(crg_tu[0])
            rce = self.cref_coord(crg_tu[2])
            return [self._wsheet.row_values(row, start_colx=rcs[1],
                                            end_colx=rce[1] + 1)
                    for row in range(rcs[0], rce[0] + 1)]
        # normal cell_ref - no check on validity of cell_ref. Leave it to xlrd
        rc_ = self.cref_coord(cell_ref)
        return self._wsheet.cell_value(*rc_)

    def _get_cell_value(self, cell, dates=datetime.date, empty=math.nan):
        """
        Returns cell value. Internal function.

        cell : xlrd.sheet.Cell
        dates : datetime.datetime or default datetime.date
        empty : str or default math.nan
            Sets empty or blank cell return value e.g 'NA', '' or nan

        """
        if cell is None:
            return None

        ctyp = cell.ctype
        value = cell.value
        # Any changes in code below also need to change cell_range_value code
        if ctyp == 1 or ctyp == 2:  # XL_CELL_TEXT=1, XL_CELL_NUMBER=2
            return value
        elif ctyp == 0 or ctyp == 6:  # XL_CELL_EMPTY=0, XL_CELL_BLANK=6,
            return empty  # NaN is easy for pandas and numpy to process
        elif ctyp == 5:  # XL_CELL_ERROR=5 return NaN
            return math.nan
        elif ctyp == 3:  # XL_CELL_DATE=3
            date_xl = xlrd.xldate.xldate_as_datetime(value, self._dmode)
            # process date options
            if dates is datetime.datetime:
                return date_xl.datetime()
            elif dates is datetime.date:
                return date_xl.date()
            return date_xl
        elif ctyp == 4:  # XL_CELL_BOOLEAN = 4
            return bool(value)
        return value

    def cell_value(self, row_no, col_no, dates=datetime.date, empty=math.nan):
        """
        Returns cell value based on sheet row number and column number.
        For cell reference use range_value.
        Use cell_range_value to return a range of cell values.
        Not efficient to use cell_value in a loop.

        row_no: int
            sheet row number 0-based
        col_no: int
            sheet column number 0-based
        dates : datetime.datetime or default datetime.date
        empty : str or default math.nan
            Sets empty or blank cell return value e.g 'NA', '' or nan

        """
        return self._get_cell_value(self._wsheet.cell(row_no, col_no),
                                    dates, empty)

    def cell_range_value(self, start_row, start_col, end_row, end_col,
                         dates=datetime.date, ndim=1, empty=math.nan):
        """
        Returns a range of cell values based on start row no, end row no and
        start column no, end column no. 0-based.
        For a single value use cell_value.
        For cell reference or range cell reference use range_value

        Parameters
        ---------
        start_row: int
            sheet start row number 0-based
        start_col: int
            sheet start column row number 0-based
        end_row: int
            sheet end row number
        end_col: int
            sheet end row number
        ndim: int
            default 1 for 1D list or 2 for 2D list
        dates : datetime.datetime or default datetime.date
        empty : str or default math.nan
            Sets empty or blank cell return value e.g 'NA', '' or nan

        Returns
        -------
        list of cell values
        """

        rtypes = [self._wsheet.row_types(row, start_colx=start_col,
                                         end_colx=end_col + 1)
                  for row in range(start_row, end_row + 1)]
        rvalues = [self._wsheet.row_values(row, start_colx=start_col,
                                           end_colx=end_col + 1)
                   for row in range(start_row, end_row + 1)]
        # Any changes in code below also need to change cell_value code
        # Modify value in range_values list to return the correct value
        # based on types
        for i in range(len(rtypes)):
            for j in range(len(rtypes[i])):
                ctyp = rtypes[i][j]
                if ctyp == 0 or ctyp == 6:  # XL_CELL_EMPTY=0, XL_CELL_BLANK=6
                    # NaN is easy for pandas and numpy to process
                    rvalues[i][j] = empty
                elif ctyp == 5:  # XL_CELL_ERROR=5 return NaN
                    rvalues[i][j] = math.nan
                elif ctyp == 3:  # XL_CELL_DATE=3
                    date_xl = xlrd.xldate.xldate_as_datetime(
                        rvalues[i][j], self._dmode)
                    # process date options
                    if dates is datetime.datetime:
                        rvalues[i][j] = date_xl
                    elif dates is datetime.date:
                        rvalues[i][j] = date_xl.date()
                    else:
                        rvalues[i][j] = date_xl
                elif ctyp == 4:  # XL_CELL_BOOLEAN = 4
                    rvalues[i][j] = bool(rvalues[i][j])
                    # no change for XL_CELL_TEXT=1, XL_CELL_NUMBER=2
        if ndim == 1:
            return [value for row in rvalues for value in row]
        return rvalues

    def set_cell_value(self, cell_ref, values):
        raise NotImplementedError

    def save(self, fname=None):
        raise NotImplementedError

    # Port Andrew Khan's java functions
    # from JExcelApi CellReferenceHelper class
    @classmethod
    def cref_row_no(cls, crs):
        """
        Get the row number of a cell reference e.g. A10 returns 9

        crs: str, cell reference

        Returns
        -------
        int, row number
        """
        numindex = cls.cref_num_index(crs)
        return int(crs[numindex:]) - 1  # pure python integer will not overflow

    # Port Andrew Khan's java functions from
    # JExcelApi CellReferenceHelper class
    @classmethod
    def cref_column_no(cls, crs):
        """
        Get the column number portion of a cell reference. 0-based.
        e.g. B25 returns 1

        crs: str, cell reference

        Returns
        -------
        int, column number
        """
        colnum = 0
        numindex = cls.cref_num_index(crs)
        str2 = crs.upper()
        start_pos = crs.rfind(cls._sheet_ind) + 1
        if crs[start_pos] == cls._fixed_ind:
            start_pos += 1

        end_pos = numindex
        if crs[numindex - 1] == cls._fixed_ind:
            end_pos -= 1

        for i in range(start_pos, end_pos):
            if i != start_pos:
                colnum = (colnum + 1) * 26
            colnum += ord(str2[i]) - ord('A')
        return colnum

    @classmethod
    def cref_coord(cls, crs):
        """
        Get the row and column number of a cell reference
        e.g. IV525 returns (524, 255)

        crs: str, cell reference

        Returns
        -------
        tuple, (row_no, col_no)
        """
        return cls.cref_row_no(crs), cls.cref_column_no(crs)

    # port from Apache poi CellReference's convertNumToColString() method
    @classmethod
    def get_col_ref(cls, col_no):
        """
        Takes in a 0-based base-10 column and returns a ALPHA-26
        representation in bytearray form.
        eg column #3 -> D

        col_no: int

        returns
        -------
        bytearray, ALPHA-26 representation in
        bytearray form. e.g. 'A' -> 65, 'B' -> 66, 'C' -> 67
        """
        # Excel counts column A as the 1st column, we treat it as the 0th one
        excel_col_num = col_no + 1

        col_ref = bytearray()
        col_remain = excel_col_num

        while col_remain > 0:
            this_part = col_remain % 26
            if this_part == 0:
                this_part = 26
            col_remain = (col_remain - this_part) // 26

            # The letter A is at 65
            col_char = (this_part + 64)  # handle as int because of bytearray
            col_ref.append(col_char)
        col_ref.reverse()
        return col_ref

    @classmethod
    def get_cell_ref(cls, row_no, col_no, rowabs=False, colabs=False):
        """
        Get the cell reference from the given row and column number

        row_no: int, Row number starts from 0
        col_no: int, Column number starts from 0
        rowabs: bool, Absolute row address if set to True. default is False
        colabs: bool, Absolute column address is set to True. default is False

        Returns
        -------
        cell reference: str
            e.g. 'A1', 'B2', 'C3', '$A1', 'B$2', '$C$3'
        """

        buf = bytearray()
        if colabs:
            buf.append(ord(cls._fixed_ind))
        # Put the column letter into the buffer
        buf.extend(cls.get_col_ref(col_no))
        if rowabs:
            buf.append(ord(cls._fixed_ind))
        # bytearray default encoding is utf-8
        buf.extend(bytearray(str(row_no + 1), 'utf-8'))
        return buf.decode()

    def row_values(self, xl_row_no: int, opt_date=datetime.date,
                   opt_empty=math.nan):
        """
        Get the cell values in the specified row number.

        Parameters
        ----------
        xl_row_no : int
            MS Excel row number. 1-based index
        opt_date : datetime.datetime or default datetime.date
        opt_empty : str or default math.nan
            Sets empty or blank cell return value e.g 'NA', '' or nan

        Returns
        -------
            list of cell values
        """
        rowv = self._wsheet.row(
            xl_row_no - 1)  # xlrd is zero-based index unlike MS Excel
        return [self._get_cell_value(cell, opt_date, opt_empty)
                for cell in rowv]

    def col_values(self, xl_col_str: str, opt_date=datetime.date,
                   opt_empty=math.nan):
        """
        Returns the cell values in the specified column address.

        Parameters
        ----------
        xl_col_str : str
            MS Excel column address e.g. 'A'
        opt_date : datetime.datetime or default datetime.date
        opt_empty : str or default math.nan
            Sets empty or blank cell return value e.g 'NA', '' or nan

        Returns
        -------
            list of cell values
        """
        col_num = self.cref_column_no(xl_col_str)  # a;ready 0-based
        colv = self._wsheet.col(col_num)
        return [self._get_cell_value(cell, opt_date, opt_empty)
                for cell in colv]


# This class is for easy migration of XLWDataLink code to preferred
# OXLDataLink or XLDataLink since it shares the same methods.
# This class uses xlwings which depends on MS Excel App.
class XLWDataLink(PyXLDatalink):
    """
    XLWDatalink uses xlwings library to extract data from
    a MS Excel workbook. Row and column numbers
    are 1-based.
    xlwings library requires MS Excel App to be installed
    otherwise it won't work.
    XLWDatalink solves some workbook errors such as:
    'xlrd.biffh.XLRDError: Workbook is encrypted' caused by
    MS Excel 95 protected file issue.
    This class is slower because it uses the MS Excel app
    therefore XLDataLink or OXLDataLink is preferred
    """

    def __init__(self, xlapp: xlwings.App, xlfname: str, xlsheet=0):
        """
        Constructor for XLWDataLink.
        Parameters
        ----------
        xlapp: xlwings.App
        xlfname: str, MS Excel file name
        xlsheet: str e.g. 'Sheet1' or int e.g. 0 for
        first sheet. Default is 0.

        Example
        -------
        >>> import xlwings as xw
        >>> from xltab import XLWDataLink
        >>> xlapp = xw.App()
        >>> xlapp.visible = False
        >>> with XLWDatalink(xlapp, 'MyExcelWb.xlsx', 'Sheet5') as xldl
        >>>     xldl.tab_form(['A3', 'C7', 'H8'])
        >>> # Close the MS Excel App to free resources
        >>> xlapp.quit()
        """
        super().__init__(xlfname, xlsheet)
        self._xlapp = xlapp  # MS Excel App
        self._xlapp.display_alerts = False
        self._wbook = self._xlapp.books.open(xlfname)
        self.sheet = xlsheet

    # Context Manager implementations
    def __enter__(self):
        return self

    # automatically close() when using 'with' statement
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    # Having problem with python destructor __del__ executing after the
    # MS Excel App is closed due to wrong timing of the
    # garbage collection routines.
    # ContextManagers implemented
    # For simplicity implement close() to explicitly close
    # the workbook like xlwings
    def close(self):
        """
        Explicitly close the workbook in the MS Excel App to
        save resources. Does not close the MS Excel App.
        Use the 'with' statement instead to avoid forgeting to close.
        """
        if self._wsheet is not None:
            self._wbook.close()
            # prevent reuse
            self._wbook = None
            self._wsheet = None

    @property
    def sheet(self):
        return self._sheet_name

    @sheet.setter
    def sheet(self, sheet_spec):
        # self._wbook.sheets[0] or  self._wbook.sheets['Sheet1']
        self._wsheet = self._wbook.sheets[sheet_spec]
        self._sheet_name = self._wsheet.name

    def __str__(self):
        return f'{self.__class__.__name__}: App = {self._xlapp},\
 Book = {self._file_name}, Sheet = {self._sheet_name}'

    def __repr__(self):
        return f'{self.__class__.__name__}({self._xlapp}, {self._file_name!r},\
 {self._sheet_name!r})'

    def app(self):
        return self._xlapp

    def range_value(self, cell_ref, dates=datetime.date, ndim=1,
                    empty=math.nan):
        if ndim == 1:
            values = self._wsheet.range(cell_ref).options(
                dates=dates, empty=empty).value
        else:
            values = self._wsheet.range(cell_ref).options(
                ndim=ndim, dates=dates, empty=empty).value
        return values

    def raw_range_value(self, cell_ref):
        """
        Retrieves the cell value or a range of cell values
        as is.
        cell_ref: str, e.g. 'G68' or 'G68:K88'
        return: cell value or list of cell values in 1-D or 2-D
        depending on how the cell reference is specified
        """
        return self._wsheet.range(cell_ref).value

    def row_values(self, xl_row_no: int, opt_date=datetime.date,
                   opt_empty=math.nan):
        """
        Not implemented for XLWDataLink because of xlwings limitations
        """
        raise NotImplementedError

    def col_values(self, xl_col_str: str, opt_date=datetime.date,
                   opt_empty=math.nan):
        """
        Not implemented for XLWDataLink because of xlwings limitations
        """
        raise NotImplementedError

    def set_cell_value(self, cell_ref, values):
        """
        Sets the cell value for a single cell or for a range of
        cells specified in the cell reference

        cell_ref: str, e.g. 'Z10' or range 'ZA1:ZC10'
        values: one cell value in any data type or values in a
        2D list with same dimension as the specificied
        cell reference range
        """
        if cell_ref is None:
            return None  # since no valid cell address do nothing
        elif isinstance(cell_ref, list):
            return None   # cell_ref should not be a list here
        elif ':' in cell_ref:  # range type of cell_ref
            rges = self._wsheet[cell_ref]  # xlwings range object
            # convert to 1-dim tuples
            rges_tu = (rge for row in rges for rge in row)
            values_tu = (value for row in values for value in row)
            tu_ls = zip(rges_tu, values_tu)
            for rge, value in tu_ls:
                rge.value = value
        else:   # single cell address
            self._wsheet.range(cell_ref).value = values

    def save(self, fname=None):
        """ Save the changes
        """
        if fname is None:
            self._wbook.save(self._file_name)
        else:
            self._wbook.save(fname)
